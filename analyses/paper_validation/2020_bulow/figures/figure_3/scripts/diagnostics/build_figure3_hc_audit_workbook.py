from __future__ import annotations

import math
import shutil
import sys


from pathlib import Path
import sys as _bootstrap_sys
from pathlib import Path as _BootstrapPath

for _candidate in _BootstrapPath(__file__).resolve().parents:
    if (_candidate / "scripts" / "plot_outputs.py").is_file():
        if str(_candidate) not in _bootstrap_sys.path:
            _bootstrap_sys.path.insert(0, str(_candidate))
        break
else:
    raise ModuleNotFoundError("Could not locate repo root containing scripts/plot_outputs.py")
from scripts.plot_outputs import REPO_ROOT

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

SCRIPT_DIR = Path(__file__).resolve().parent
FIGURE_DIR = SCRIPT_DIR.parent
ANALYSIS_ROOT = FIGURE_DIR.parents[2]

if str(ANALYSIS_ROOT) not in sys.path:
    sys.path.insert(0, str(ANALYSIS_ROOT))
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from scripts._env import require_epcsaft_install

require_epcsaft_install()

from epcsaft.parameters import get_prop_dict
from scripts._epcsaft_oop import epcsaft_density, epcsaft_fugacity_coefficient_terms, epcsaft_pressure
import _plot_common as common

R_GAS = 8.31446261815324
N_AV = 6.022140857e23
PI = math.pi
T_REF = 298.15
P_REF = 1.0e5
EPS = 1.0e-8
EPS_INF = 1.0e-12
IONS = ("Li+", "Na+", "K+", "F-", "Cl-", "Br-", "I-")
SOURCE_WORKBOOK = Path(
    r"C:\Users\Tanner\Documents\Learning Resources\Classes\0_Spring_2024\ePC-SAFT Calculations - Part 1 - Helmholtz Free Energy.xlsm"
)
OUTPUT_WORKBOOK = REPO_ROOT / "workbooks" / "figure3_hc_audit.xlsm"
DATA_PATH = common.analysis_data_path(
    ANALYSIS_ROOT / "figure_3", "water_contributions.csv", kind="processed", category="figure_3"
)


def _species_for_ion(ion: str) -> list[str]:
    if ion in {"Li+", "Na+", "K+"}:
        return [ion, "Cl-", "Water"]
    if ion == "F-":
        return ["Na+", "F-", "Water"]
    if ion in {"Cl-", "Br-", "I-"}:
        return ["Na+", ion, "Water"]
    raise KeyError(ion)


def _effective_diameter(sigma: float, epsilon: float, charge: float, mode: int) -> float:
    if abs(charge) <= 1.0e-12:
        return sigma * (1.0 - 0.12 * math.exp(-3.0 * epsilon / T_REF))
    if mode == 0:
        return sigma
    if mode == 1:
        return sigma * (1.0 - 0.12)
    if mode == 2:
        return sigma * (1.0 - 0.12 * math.exp(-3.0 * epsilon / T_REF))
    raise ValueError(mode)


def _infinite_dilution_state(ion: str, d_ion_mode: int) -> dict[str, object]:
    species = _species_for_ion(ion)
    x = np.asarray([EPS, EPS, 1.0 - 2.0 * EPS], dtype=float)
    user_options = {"elec_model": {"DH_model": {"d_ion_mode": d_ion_mode}}}
    params = get_prop_dict("2020_Bulow", species, x, T_REF, user_options=user_options)
    rho = epcsaft_density(T_REF, P_REF, x, params, phase="liq")

    z = np.asarray(params.get("z", []), dtype=float)
    idx_ion = np.where(np.abs(z) > 1.0e-12)[0]
    idx_solv = np.where(np.abs(z) <= 1.0e-12)[0]

    x_ref = x.copy()
    x_ref[idx_ion] = 0.0
    x_ref[idx_solv] = x_ref[idx_solv] / float(np.sum(x_ref[idx_solv]))
    p_ref = epcsaft_pressure(T_REF, rho, x_ref, params)

    x_inf = x_ref.copy()
    ion_idx = species.index(ion)
    x_inf[ion_idx] = EPS_INF
    x_inf /= float(np.sum(x_inf))
    rho_inf = epcsaft_density(T_REF, p_ref, x_inf, params, phase="liq")

    terms = epcsaft_fugacity_coefficient_terms(T_REF, rho_inf, x_inf, params)
    m = np.asarray(params["m"], dtype=float)
    s = np.asarray(params["s"], dtype=float)
    e = np.asarray(params["e"], dtype=float)
    d = np.asarray([_effective_diameter(s[i], e[i], z[i], d_ion_mode) for i in range(len(species))], dtype=float)
    den = rho_inf * N_AV / 1.0e30
    zeta = np.asarray(
        [PI / 6.0 * den * np.sum(x_inf * m * np.power(d, n)) for n in range(4)],
        dtype=float,
    )
    i = ion_idx
    ghs_ii = (
        1.0 / (1.0 - zeta[3])
        + (d[i] / 2.0) * 3.0 * zeta[2] / (1.0 - zeta[3]) ** 2
        + (d[i] / 2.0) ** 2 * 2.0 * zeta[2] ** 2 / (1.0 - zeta[3]) ** 3
    )
    a_hs = (
        1.0
        / zeta[0]
        * (
            3.0 * zeta[1] * zeta[2] / (1.0 - zeta[3])
            + zeta[2] ** 3 / (zeta[3] * (1.0 - zeta[3]) ** 2)
            + ((zeta[2] ** 3 / zeta[3] ** 2) - zeta[0]) * math.log(1.0 - zeta[3])
        )
    )
    return {
        "species": species,
        "params": params,
        "x_inf": x_inf,
        "rho_inf": float(rho_inf),
        "p_ref": float(p_ref),
        "ion_index": ion_idx,
        "d": d,
        "zeta": zeta,
        "ghs_ii": float(ghs_ii),
        "a_hs": float(a_hs),
        "terms": terms,
    }


def _write_active_mode_summary(ws, start_row: int, frame) -> int:
    headers = [
        "Ion",
        "Paper hc",
        "mode1 hc",
        "mode1 - paper",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    row = start_row + 1
    for ion in IONS:
        paper = float(frame.scalar("hc avg", ion))
        state = _infinite_dilution_state(ion, 1)
        idx = int(state["ion_index"])
        mu_hc = float(R_GAS * T_REF * np.asarray(state["terms"]["mu_hc"], dtype=float)[idx] / 1000.0)
        row_values = [ion, paper, mu_hc, mu_hc - paper]
        for col, value in enumerate(row_values, start=1):
            ws.cell(row, col, value)
        row += 1
    return row


def _write_active_mode_detail(ws, start_row: int, frame) -> int:
    headers = [
        "Ion",
        "Species order",
        "x_inf",
        "m_ion",
        "sigma_ion",
        "epsilon_ion",
        "d_ion(mode1)",
        "zeta0",
        "zeta1",
        "zeta2",
        "zeta3",
        "ghs_ii",
        "excel-style a_hs",
        "excel-style a_hc",
        "package a_hc",
        "delta a_hc",
        "dadx_hc(i)",
        "sum_x_dadx_hc",
        "z_raw_hc",
        "excel-style mu_hc",
        "mu_hc",
        "delta mu_hc",
        "lnfug_hc",
        "paper hc",
        "mu_hc - paper",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8F3E8")
    row = start_row + 1
    for ion in IONS:
        state = _infinite_dilution_state(ion, 1)
        idx = int(state["ion_index"])
        terms = state["terms"]
        params = state["params"]
        m = np.asarray(params["m"], dtype=float)
        s = np.asarray(params["s"], dtype=float)
        e = np.asarray(params["e"], dtype=float)
        x_inf = np.asarray(state["x_inf"], dtype=float)
        # Use the exact code-level diagonal g_hs values for all species.
        ghs_diag = np.empty(len(m), dtype=float)
        zeta = np.asarray(state["zeta"], dtype=float)
        d = np.asarray(state["d"], dtype=float)
        for comp_idx in range(len(m)):
            ghs_diag[comp_idx] = (
                1.0 / (1.0 - zeta[3])
                + (d[comp_idx] / 2.0) * 3.0 * zeta[2] / (1.0 - zeta[3]) ** 2
                + (d[comp_idx] / 2.0) ** 2 * 2.0 * zeta[2] ** 2 / (1.0 - zeta[3]) ** 3
            )
        excel_a_hc = float(state["a_hs"] * float(np.sum(x_inf * m)) - np.sum(x_inf * (m - 1.0) * np.log(ghs_diag)))
        mu_hc = float(R_GAS * T_REF * np.asarray(terms["mu_hc"], dtype=float)[idx] / 1000.0)
        excel_mu_hc = float(
            R_GAS
            * T_REF
            * (
                float(terms["a_hc"])
                + float(terms["z_raw_hc"])
                + float(np.asarray(terms["dadx_hc"], dtype=float)[idx])
                - float(terms["sum_x_dadx_hc"])
            )
            / 1000.0
        )
        lnfug_hc = float(R_GAS * T_REF * np.asarray(terms["lnfugcoef_hc"], dtype=float)[idx] / 1000.0)
        row_values = [
            ion,
            ", ".join(state["species"]),
            ", ".join(f"{float(v):.3e}" for v in np.asarray(state["x_inf"], dtype=float)),
            float(m[idx]),
            float(s[idx]),
            float(e[idx]),
            float(np.asarray(state["d"], dtype=float)[idx]),
            float(np.asarray(state["zeta"], dtype=float)[0]),
            float(np.asarray(state["zeta"], dtype=float)[1]),
            float(np.asarray(state["zeta"], dtype=float)[2]),
            float(np.asarray(state["zeta"], dtype=float)[3]),
            float(state["ghs_ii"]),
            float(state["a_hs"]),
            excel_a_hc,
            float(terms["a_hc"]),
            excel_a_hc - float(terms["a_hc"]),
            float(R_GAS * T_REF * np.asarray(terms["dadx_hc"], dtype=float)[idx] / 1000.0),
            float(R_GAS * T_REF * float(terms["sum_x_dadx_hc"]) / 1000.0),
            float(terms["z_raw_hc"]),
            excel_mu_hc,
            mu_hc,
            excel_mu_hc - mu_hc,
            lnfug_hc,
            float(frame.scalar("hc avg", ion)),
            mu_hc - float(frame.scalar("hc avg", ion)),
        ]
        for col, value in enumerate(row_values, start=1):
            ws.cell(row, col, value)
        row += 1
    return row


def build_workbook() -> Path:
    frame = common.load_indexed_csv(DATA_PATH)
    OUTPUT_WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SOURCE_WORKBOOK, OUTPUT_WORKBOOK)
    wb = load_workbook(OUTPUT_WORKBOOK, keep_vba=True)
    if "Figure3 HC Audit" in wb.sheetnames:
        del wb["Figure3 HC Audit"]
    ws = wb.create_sheet("Figure3 HC Audit")

    ws["A1"] = "Figure 3 Hard-Chain Audit"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Source workbook copied from:"
    ws["B2"] = str(SOURCE_WORKBOOK)
    ws["A3"] = "Audit basis:"
    ws["B3"] = "Infinite-dilution water hydration path used by the 2020 Figure 3 scripts"
    ws["A4"] = "Notes:"
    ws["B4"] = (
        "This workbook keeps only the active 2020 setting d_ion_mode = 1. The detail section compares workbook-style hard-chain equations against package values at the same infinite-dilution state."
    )
    ws["A2"].font = ws["A3"].font = ws["A4"].font = Font(bold=True)
    ws["B3"].alignment = Alignment(wrap_text=True)
    ws["B4"].alignment = Alignment(wrap_text=True)

    next_row = _write_active_mode_summary(ws, 7, frame)
    next_row += 2
    _write_active_mode_detail(ws, next_row, frame)

    widths = {
        "A": 12,
        "B": 28,
        "C": 34,
        "D": 11,
        "E": 11,
        "F": 11,
        "G": 13,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 12,
        "N": 12,
        "O": 14,
        "P": 16,
        "Q": 12,
        "R": 12,
        "S": 12,
        "T": 12,
        "U": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(OUTPUT_WORKBOOK)
    wb.close()
    return OUTPUT_WORKBOOK


if __name__ == "__main__":
    path = build_workbook()
    print(path)

