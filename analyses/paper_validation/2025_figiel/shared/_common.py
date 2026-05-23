from __future__ import annotations

import csv
import json
import math
import sys


from collections import defaultdict
from functools import lru_cache
from pathlib import Path
import sys as _bootstrap_sys
from pathlib import Path as _BootstrapPath

for _candidate in _BootstrapPath(__file__).resolve().parents:
    if (_candidate / "scripts" / "plot_outputs.py").is_file():
        if str(_candidate) not in _bootstrap_sys.path:
            _bootstrap_sys.path.insert(0, str(_candidate))
        break
else:
    raise ModuleNotFoundError("Could not locate repo root containing scripts/plot_outputs.py")
from scripts.plot_outputs import REPO_ROOT
from typing import Dict, Iterable, List, Tuple

import matplotlib
import numpy as np

matplotlib.use("Agg")
import matplotlib.pyplot as plt

if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))
SRC_ROOT = REPO_ROOT / "src"
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))

from scripts._env import require_epcsaft_install
from scripts.plot_outputs import figure_output_path, save_plot_figure

require_epcsaft_install()

from epcsaft.parameters import get_prop_dict
from scripts._epcsaft_oop import as_mixture

T_REF = 298.15
P_REF = 1.0e5

SOLVENT_MW = {
    "water": 18.01528e-3,
    "methanol": 32.04e-3,
    "ethanol": 46.068e-3,
}

SALT_SPECS = {
    "HCl": {"cation": "H+", "anion": "Cl-", "z_cation": 1, "z_anion": -1},
    "LiBr": {"cation": "Li+", "anion": "Br-", "z_cation": 1, "z_anion": -1},
    "LiCl": {"cation": "Li+", "anion": "Cl-", "z_cation": 1, "z_anion": -1},
    "LiI": {"cation": "Li+", "anion": "I-", "z_cation": 1, "z_anion": -1},
    "NaBr": {"cation": "Na+", "anion": "Br-", "z_cation": 1, "z_anion": -1},
    "NaCl": {"cation": "Na+", "anion": "Cl-", "z_cation": 1, "z_anion": -1},
    "NaI": {"cation": "Na+", "anion": "I-", "z_cation": 1, "z_anion": -1},
    "KCl": {"cation": "K+", "anion": "Cl-", "z_cation": 1, "z_anion": -1},
    "KBr": {"cation": "K+", "anion": "Br-", "z_cation": 1, "z_anion": -1},
    "KI": {"cation": "K+", "anion": "I-", "z_cation": 1, "z_anion": -1},
}

ION_TO_REFERENCE_SALT = {
    "H+": "HCl",
    "Li+": "LiCl",
    "Na+": "NaCl",
    "K+": "KCl",
    "Cl-": "NaCl",
    "Br-": "NaBr",
    "I-": "NaI",
}

ORGANIC_COLOR = "#d9891b"
GREEN_COLOR = "#228b22"
GRAY_COLOR = "0.45"
LIGHT_GRAY = "0.82"
BROWN_COLOR = "#7a1f1f"
BLUE_COLOR = "#1f4da8"


def _freeze_comp(comp: Dict[str, float]) -> Tuple[Tuple[str, float], ...]:
    return tuple(sorted((str(k), round(float(v), 12)) for k, v in comp.items()))


def _thaw_comp(comp_key: Tuple[Tuple[str, float], ...]) -> Dict[str, float]:
    return {str(k): float(v) for k, v in comp_key}


def _freeze_user_options(user_options: dict | None) -> str:
    return json.dumps(user_options or {}, sort_keys=True, separators=(",", ":"))


def _thaw_user_options(user_options_key: str) -> dict | None:
    data = json.loads(user_options_key)
    return data if data else None


def configure_style() -> None:
    plt.rcParams.update(
        {
            "font.size": 10,
            "font.family": "DejaVu Serif",
            "axes.linewidth": 1.0,
            "axes.grid": False,
            "xtick.direction": "in",
            "ytick.direction": "in",
            "xtick.top": True,
            "ytick.right": True,
            "legend.frameon": False,
            "mathtext.default": "regular",
        }
    )


def panel_label(ax, label: str) -> None:
    ax.text(0.03, 0.96, label, transform=ax.transAxes, ha="left", va="top", fontsize=11)


def save_figure(fig, path: Path) -> None:
    path = figure_output_path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    save_plot_figure(fig, path, dpi=300)


def save_panel_figure(
    plotter,
    path: Path,
    figsize: tuple[float, float] = (3.6, 3.1),
    legend_handles=None,
    legend_labels=None,
    legend_kwargs: dict | None = None,
) -> None:
    configure_style()
    fig, ax = plt.subplots(figsize=figsize)
    plotter(ax)
    if legend_handles and legend_labels:
        kwargs = {"loc": "best", "fontsize": 8, "frameon": False}
        if legend_kwargs:
            kwargs.update(legend_kwargs)
        ax.legend(legend_handles, legend_labels, **kwargs)
    save_figure(fig, path)
    plt.close(fig)


def parse_float(value) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        val = float(text)
    except ValueError:
        return None
    if not math.isfinite(val):
        return None
    return float(val)


def read_csv_rows(path: Path) -> Tuple[List[str], List[Dict[str, str]]]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        fields = [h.strip() for h in (reader.fieldnames or []) if h and h.strip()]
        rows: List[Dict[str, str]] = []
        for row in reader:
            clean: Dict[str, str] = {}
            for key, value in row.items():
                if not key:
                    continue
                ks = key.strip()
                if not ks:
                    continue
                clean[ks] = value.strip() if isinstance(value, str) else value
            rows.append(clean)
    return fields, rows


def write_xlsx_to_csv(xlsx_path: Path, csv_path: Path | None = None) -> Path:
    from openpyxl import load_workbook

    if csv_path is None:
        csv_path = xlsx_path.with_suffix(".csv")
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(list(row))
    return csv_path


def species_for_combo(salt: str, solvent_system: str) -> List[str]:
    salt_spec = SALT_SPECS[salt]
    solvents = [s for s in solvent_system.split("-") if s]
    solvent_species: List[str] = []
    for solvent in solvents:
        if solvent == "water":
            solvent_species.append("H2O-2B-Li" if salt.startswith("Li") else "H2O-2B-NaCl")
        elif solvent == "methanol":
            solvent_species.append("Methanol")
        elif solvent == "ethanol":
            solvent_species.append("Ethanol")
        else:
            raise ValueError(f"Unsupported solvent '{solvent}'.")
    return [salt_spec["cation"], salt_spec["anion"], *solvent_species]


def normalized_comp(solvent_system: str, comp: Dict[str, float]) -> Dict[str, float]:
    solvents = [s for s in solvent_system.split("-") if s]
    if len(solvents) == 1:
        return {solvents[0]: 1.0}
    vals = {s: float(comp.get(s, 0.0)) for s in solvents}
    total = sum(vals.values())
    if total <= 0.0:
        return {s: 1.0 / len(solvents) for s in solvents}
    return {s: vals[s] / total for s in solvents}


def mw_mix(solvent_system: str, comp: Dict[str, float]) -> float:
    frac = normalized_comp(solvent_system, comp)
    return sum(frac[s] * SOLVENT_MW[s] for s in frac)


def stoich_for_salt(salt: str) -> Tuple[int, int]:
    spec = SALT_SPECS[salt]
    zc = abs(int(round(spec["z_cation"])))
    za = abs(int(round(spec["z_anion"])))
    g = math.gcd(zc, za)
    return za // g, zc // g


def salt_mole_fraction_from_molality(molality: np.ndarray, solvent_system: str, comp: Dict[str, float]) -> np.ndarray:
    m = np.asarray(molality, dtype=float)
    n_solv = 1.0 / mw_mix(solvent_system, comp)
    denom = n_solv + m
    return np.where(denom > 0.0, m / denom, 0.0)


def molality_to_species_molefraction(
    molality: float, salt: str, solvent_system: str, comp: Dict[str, float]
) -> np.ndarray:
    species = species_for_combo(salt, solvent_system)
    solvents = [s for s in solvent_system.split("-") if s]
    solvent_species = species[2:]
    frac = normalized_comp(solvent_system, comp)
    mw = mw_mix(solvent_system, frac)
    n_solv_total = 1.0 / mw
    nu_cat, nu_an = stoich_for_salt(salt)
    n_totals: Dict[str, float] = {sp: 0.0 for sp in species}
    for solvent_key, sp_name in zip(solvents, solvent_species):
        n_totals[sp_name] += frac[solvent_key] * n_solv_total
    n_totals[species[0]] += nu_cat * molality
    n_totals[species[1]] += nu_an * molality
    total = sum(n_totals.values())
    if total <= 0.0:
        raise ValueError("Non-positive total moles in mole-fraction conversion.")
    return np.asarray([n_totals[sp] / total for sp in species], dtype=float)


def build_params(
    dataset_name: str, salt: str, solvent_system: str, comp: Dict[str, float], user_options: dict | None = None
) -> dict:
    x_ref = molality_to_species_molefraction(1e-8, salt, solvent_system, comp)
    return get_prop_dict(dataset_name, species_for_combo(salt, solvent_system), x_ref, T_REF, user_options=user_options)


def pair_key_for_salt(salt: str) -> str:
    spec = SALT_SPECS[salt]
    return f"{spec['cation']}{spec['anion']}"


def mean_ionic_activity_curve(
    dataset_name: str,
    salt: str,
    solvent_system: str,
    comp: Dict[str, float],
    m_max: float,
    points: int = 500,
    user_options: dict | None = None,
) -> Tuple[np.ndarray, np.ndarray]:
    comp_key = _freeze_comp(comp)
    user_options_key = _freeze_user_options(user_options)
    grid, gamma = _mean_ionic_activity_curve_cached(
        dataset_name,
        salt,
        solvent_system,
        comp_key,
        float(m_max),
        int(points),
        user_options_key,
    )
    return np.asarray(grid, dtype=float).copy(), np.asarray(gamma, dtype=float).copy()


@lru_cache(maxsize=256)
def _mean_ionic_activity_curve_cached(
    dataset_name: str,
    salt: str,
    solvent_system: str,
    comp_key: Tuple[Tuple[str, float], ...],
    m_max: float,
    points: int,
    user_options_key: str,
) -> Tuple[np.ndarray, np.ndarray]:
    comp = _thaw_comp(comp_key)
    user_options = _thaw_user_options(user_options_key)
    grid = np.linspace(0.0, float(m_max), int(points))
    params = build_params(dataset_name, salt, solvent_system, comp, user_options=user_options)
    species = species_for_combo(salt, solvent_system)
    pair_key = pair_key_for_salt(salt)
    mixture = as_mixture(params, species=species)
    gamma = np.empty_like(grid)
    for idx, m in enumerate(grid):
        m_eval = max(float(m), 1e-12)
        x = molality_to_species_molefraction(m_eval, salt, solvent_system, comp)
        state = mixture.state(T=T_REF, x=x, P=P_REF, phase="liq")
        gamma[idx] = state.activity_coefficient(species=species, mean_ionic_form=True, basis="molality")[pair_key]
    return grid, gamma


def extract_comp(row: Dict[str, str], solvent_system: str) -> Dict[str, float]:
    solvents = [s for s in solvent_system.split("-") if s]
    if len(solvents) == 1:
        return {solvents[0]: 1.0}
    comp: Dict[str, float] = {}
    mapping = {
        "x_h2o": "water",
        "x_water": "water",
        "x_methanol": "methanol",
        "x_meoh": "methanol",
        "x_ethanol": "ethanol",
        "x_etoh": "ethanol",
    }
    for key, value in row.items():
        k = key.strip().lower()
        if k in mapping and mapping[k] in solvents:
            parsed = parse_float(value)
            if parsed is not None:
                comp[mapping[k]] = parsed
    if len(solvents) == 2 and len(comp) == 1:
        known = next(iter(comp))
        other = [s for s in solvents if s != known][0]
        comp[other] = 1.0 - comp[known]
    return normalized_comp(solvent_system, comp)


def comp_signature(solvent_system: str, comp: Dict[str, float]) -> Tuple[Tuple[str, float], ...]:
    solvents = [s for s in solvent_system.split("-") if s]
    if len(solvents) <= 1:
        return tuple()
    return tuple((s, round(float(comp.get(s, 0.0)), 6)) for s in solvents)


def read_miac_dataset(path: Path, solvent_system: str) -> List[Dict[str, object]]:
    fields, rows = read_csv_rows(path)
    lookup = {field.lower(): field for field in fields}
    molality_key = next((lookup[c] for c in ("molality", "m") if c in lookup), None)
    miac_m_key = next((lookup[c] for c in ("miac_m", "gamma") if c in lookup), None)
    miac_key = next((lookup[c] for c in ("miac", "y") if c in lookup), None)
    x_key = lookup.get("mole_fraction")
    if molality_key is None or miac_m_key is None:
        raise ValueError(f"Missing required columns in {path}.")
    data: List[Dict[str, object]] = []
    for row in rows:
        m = parse_float(row.get(molality_key))
        gm = parse_float(row.get(miac_m_key))
        if m is None or gm is None:
            continue
        comp = extract_comp(row, solvent_system)
        data.append(
            {
                "molality": m,
                "miac_m": gm,
                "miac": parse_float(row.get(miac_key)) if miac_key else None,
                "mole_fraction": parse_float(row.get(x_key)) if x_key else None,
                "comp": comp,
                "signature": comp_signature(solvent_system, comp),
                "source": str(row.get(lookup.get("source", ""), "") or "").strip(),
            }
        )
    data.sort(key=lambda entry: float(entry["molality"]))
    return data


def group_by_signature(
    entries: List[Dict[str, object]],
) -> Dict[Tuple[Tuple[str, float], ...], List[Dict[str, object]]]:
    grouped: Dict[Tuple[Tuple[str, float], ...], List[Dict[str, object]]] = defaultdict(list)
    for entry in entries:
        grouped[entry["signature"]].append(entry)
    for rows in grouped.values():
        rows.sort(key=lambda entry: float(entry["molality"]))
    return dict(grouped)


def organic_weight_fraction(solvent_system: str, comp: Dict[str, float]) -> float:
    solvents = [s for s in solvent_system.split("-") if s]
    if len(solvents) != 2 or "water" not in solvents:
        raise ValueError("Organic weight fraction only implemented for water-organic binaries.")
    organic = [s for s in solvents if s != "water"][0]
    frac = normalized_comp(solvent_system, comp)
    num = frac[organic] * SOLVENT_MW[organic]
    den = num + frac["water"] * SOLVENT_MW["water"]
    return num / den if den > 0.0 else float("nan")


def target_weight_fraction_to_comp(solvent_system: str, target_w_org: float) -> Dict[str, float]:
    solvents = [s for s in solvent_system.split("-") if s]
    if len(solvents) != 2 or "water" not in solvents:
        raise ValueError("Target weight fraction conversion only implemented for water-organic binaries.")
    organic = [s for s in solvents if s != "water"][0]
    w = float(target_w_org)
    mw_o = SOLVENT_MW[organic]
    mw_w = SOLVENT_MW["water"]
    x_org = (w / mw_o) / ((w / mw_o) + ((1.0 - w) / mw_w))
    return {"water": 1.0 - x_org, organic: x_org}


def closest_group_to_weight_fraction(
    entries: List[Dict[str, object]], solvent_system: str, target_w_org: float
) -> Tuple[List[Dict[str, object]], Dict[str, float], float]:
    grouped = group_by_signature(entries)
    candidates: List[Tuple[float, List[Dict[str, object]], Dict[str, float], float]] = []
    for rows in grouped.values():
        comp = dict(rows[0]["comp"])
        w_org = organic_weight_fraction(solvent_system, comp)
        candidates.append((abs(w_org - target_w_org), rows, comp, w_org))
    if not candidates:
        raise ValueError("No composition groups available.")
    candidates.sort(key=lambda item: item[0])
    _, rows, comp, w_org = candidates[0]
    return rows, comp, w_org


def safe_label_for_weight_fraction(w_org: float) -> str:
    return f"$w_{{org}}^{{salt-free}} = {w_org:.1f}$"


def gsolv_ion(
    dataset_name: str,
    ion: str,
    solvent_system: str,
    comp: Dict[str, float],
    user_options: dict | None = None,
) -> float:
    return _gsolv_ion_cached(
        dataset_name,
        ion,
        solvent_system,
        _freeze_comp(comp),
        _freeze_user_options(user_options),
    )


@lru_cache(maxsize=2048)
def _gsolv_ion_cached(
    dataset_name: str,
    ion: str,
    solvent_system: str,
    comp_key: Tuple[Tuple[str, float], ...],
    user_options_key: str,
) -> float:
    comp = _thaw_comp(comp_key)
    user_options = _thaw_user_options(user_options_key)
    salt = ION_TO_REFERENCE_SALT[ion]
    species = species_for_combo(salt, solvent_system)
    x = molality_to_species_molefraction(1e-8, salt, solvent_system, comp)
    params = get_prop_dict(dataset_name, species, x, T_REF, user_options=user_options)
    mixture = as_mixture(params, species=species)
    state = mixture.state(T=T_REF, x=x, P=P_REF, phase="liq")
    values = state.solvation_free_energy(species=species)
    if ion not in values:
        raise KeyError(f"Ion '{ion}' not returned from epcsaft_solvation_free_energy for {species}.")
    return float(values[ion]) / 1000.0


def transfer_curve(
    dataset_name: str,
    ion: str,
    organic_solvent: str,
    x_org_grid: np.ndarray,
    user_options: dict | None = None,
) -> np.ndarray:
    x_org_key = tuple(round(float(v), 12) for v in np.asarray(x_org_grid, dtype=float).tolist())
    out = _transfer_curve_cached(
        dataset_name,
        ion,
        organic_solvent,
        x_org_key,
        _freeze_user_options(user_options),
    )
    return np.asarray(out, dtype=float).copy()


@lru_cache(maxsize=256)
def _transfer_curve_cached(
    dataset_name: str,
    ion: str,
    organic_solvent: str,
    x_org_key: Tuple[float, ...],
    user_options_key: str,
) -> np.ndarray:
    user_options = _thaw_user_options(user_options_key)
    x_org_arr = np.asarray(x_org_key, dtype=float)
    water_ref = gsolv_ion(dataset_name, ion, "water", {"water": 1.0}, user_options=user_options)
    out = np.empty_like(x_org_arr)
    solvent_system = f"water-{organic_solvent}"
    for idx, x_org in enumerate(x_org_arr):
        comp = {"water": 1.0 - float(x_org), organic_solvent: float(x_org)}
        out[idx] = gsolv_ion(dataset_name, ion, solvent_system, comp, user_options=user_options) - water_ref
    return out


def literature_gsolv_water() -> Dict[str, float]:
    path = REPO_ROOT / "data" / "G_solv" / "water.csv"
    _, rows = read_csv_rows(path)
    out: Dict[str, float] = {}
    for row in rows:
        ion = str(row.get("Ion", "")).strip()
        value = parse_float(row.get("Gsolv (kJ/mol)"))
        if ion and value is not None:
            out[f"{ion}+" if ion in {"H", "Li", "Na", "K"} else f"{ion}-" if ion in {"Cl", "Br", "I"} else ion] = value
    return out
